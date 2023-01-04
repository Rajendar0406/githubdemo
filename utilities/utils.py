import softest
from openpyxl import Workbook,load_workbook
import csv

class Utils(softest.TestCase):
    def assertListitemtext(self, list, value):
        for Arrival in list:
            print("The text is:" ,Arrival.text)
            self.soft_assert(self.assertEqual, Arrival.text, value)
            if Arrival.text == value:
                print("test passed")
            else:
                print("test failed")

        self.assert_all()

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        datalist = []
        csvdata = open(filename, "r")
        reader = csv.reader(csvdata)
        next(reader)
        for rows in reader:
            datalist.append(rows)
        return datalist



